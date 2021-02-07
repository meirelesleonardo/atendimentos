import sys
import os
from PyQt5.QtWidgets import  QMainWindow, QApplication, QTableWidgetItem, QFileDialog, QMessageBox, QComboBox
from PyQt5 import uic
from PyQt5.QtCore import  QDate, QDateTime
from PyQt5 import QtCore, QtGui, QtWidgets
import datetime
import openpyxl
import ConectarSqlite
import getpass
import agenda
from editAtendimento import Editar



global setor


class Tecnico():
    def __init__(self, mat_gm, user, nome, setor, senha):
        self.mat = mat_gm
        self.nome = nome
        self.user = user
        self.password = senha
        self.setor = setor
    def setMat(self, mat):
        self.mat = mat
    def setNome(self, nome):
        self.nome = nome
    def setUser(self, user):
        self.user = user
    def setSenha(self, password):
        self.password = password
    def setSetor(self, setor):
        self.setor = setor
    def getMat(self):
        return self.mat_gm
    def getUser(self):
        return self.user
    def getNome(self):
        return self.nome
    def getSetor(self):
        return self.setor
    def getSenha(self):
        return self.senha

def acesso(matT,cred):
    sql = ("SELECT acesso FROM tecnicos WHERE mat_gm = '"+matT+"';")
    result = (ConectarSqlite.conectarBd(sql))

    acesso = result[0][0]
    n = acesso.count(cred)

    if n > 0:
        return True
    else:
        return False

def matPadrao(mat):
    #mat = self.matTecnico.text()
    mat1 = mat
    try:
        if mat[0] == "6":
            if len(mat) == 9:
                return mat
            try:
                if mat[-2] != "-":
                    ajust = mat[0:-1] + '-'+mat[-1]
                    if len(ajust) == 9:
                        return ajust
                if mat[4] != '.':
                    if  mat[-2] != "-":
                        ajust2 = mat[0:3] + '.'+ajust[3:]
                        if len(ajust2) == 9:
                            return ajust2
                if mat[-2] == "-":
                    ajust3 = mat[0:3] + '.' + mat[3:]
                    if len(ajust3) == 9:
                        return ajust3
            except:
                return mat1
        else:
            return mat1
    except:
        return mat1

def nomeT(mat):
    #Gsu.matPadrao(self)

    try:
        matT = mat
        sql = ("SELECT nome FROM tecnicos WHERE mat_gm = '"+matT+"';")
        nome = (ConectarSqlite.conectarBd(sql))
        return nome[0][0]
        #self.lbNomeT.setText(nome[0][0])
    except:
        return mat

class EditarTipoAtendimento(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        # Carregando EditarServico
        #uic.loadUi("EditarServico.ui", self)
    ######### Inicio cópia .ui ################
        self.setObjectName("wEditar")
        self.resize(585, 700)
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.btAddTipo = QtWidgets.QPushButton(self.centralwidget)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("IMAGENS/Add.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.btAddTipo.setIcon(icon)
        self.btAddTipo.setObjectName("btAddTipo")
        self.gridLayout.addWidget(self.btAddTipo, 1, 0, 1, 1)
        self.btRemoverTipo = QtWidgets.QPushButton(self.centralwidget)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("IMAGENS/X.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.btRemoverTipo.setIcon(icon1)
        self.btRemoverTipo.setObjectName("btRemoverTipo")
        self.gridLayout.addWidget(self.btRemoverTipo, 1, 1, 1, 1)
        self.btSalvarTipo = QtWidgets.QPushButton(self.centralwidget)
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("IMAGENS/Save-icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.btSalvarTipo.setIcon(icon2)
        self.btSalvarTipo.setObjectName("btSalvarTipo")
        self.gridLayout.addWidget(self.btSalvarTipo, 1, 2, 1, 1)
        self.btCancelarTipo = QtWidgets.QPushButton(self.centralwidget)
        self.btCancelarTipo.setObjectName("btCancelarTipo")
        self.gridLayout.addWidget(self.btCancelarTipo, 1, 3, 1, 1)
        self.scrollArea = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents_2 = QtWidgets.QWidget()
        self.scrollAreaWidgetContents_2.setGeometry(QtCore.QRect(0, 0, 565, 656))
        self.scrollAreaWidgetContents_2.setObjectName("scrollAreaWidgetContents_2")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents_2)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.tableEditarServico = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
        self.tableEditarServico.setLineWidth(1)
        self.tableEditarServico.setShowGrid(True)
        self.tableEditarServico.setWordWrap(True)
        self.tableEditarServico.setRowCount(0)
        self.tableEditarServico.setColumnCount(3)
        self.tableEditarServico.setObjectName("tableEditarServico")
        item = QtWidgets.QTableWidgetItem()
        self.tableEditarServico.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableEditarServico.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableEditarServico.setHorizontalHeaderItem(2, item)
        self.tableEditarServico.horizontalHeader().setVisible(True)
        self.tableEditarServico.horizontalHeader().setHighlightSections(True)
        self.tableEditarServico.verticalHeader().setVisible(True)
        self.gridLayout_5.addWidget(self.tableEditarServico, 0, 0, 1, 1)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents_2)
        self.gridLayout.addWidget(self.scrollArea, 0, 0, 1, 4)
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 585, 21))
        self.menubar.setObjectName("menubar")
        self.menuEditar_Dados = QtWidgets.QMenu(self.menubar)
        self.menuEditar_Dados.setObjectName("menuEditar_Dados")
        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.menubar.addAction(self.menuEditar_Dados.menuAction())

        self.retranslateUi(self)
        self.btCancelarTipo.clicked.connect(self.close)
        QtCore.QMetaObject.connectSlotsByName(self)

    def retranslateUi(self, wEditar):
        _translate = QtCore.QCoreApplication.translate
        wEditar.setWindowTitle(_translate("wEditar", "DDT_GSU"))
        self.btAddTipo.setText(_translate("wEditar", "Add"))
        self.btRemoverTipo.setText(_translate("wEditar", "Remover"))
        self.btSalvarTipo.setText(_translate("wEditar", "Salvar Alterações"))
        self.btCancelarTipo.setText(_translate("wEditar", "Cancelar"))
        item = self.tableEditarServico.horizontalHeaderItem(0)
        item.setText(_translate("wEditar", "TIPO DE SERVIÇO"))
        item = self.tableEditarServico.horizontalHeaderItem(1)
        item.setText(_translate("wEditar", "GRUPO"))
        item = self.tableEditarServico.horizontalHeaderItem(2)
        item.setText(_translate("wEditar", "ID"))
        self.menuEditar_Dados.setTitle(_translate("wEditar", "Editar Dados"))
    ######### Fim cópia .ui ####################
        self.tableEditarServico.setColumnWidth(0, 250)
        self.tableEditarServico.setColumnWidth(1, 250)
        self.tableEditarServico.setColumnWidth(2, 0)
        self.btSalvarTipo.clicked.connect(self.salvarTipo)
        self.btAddTipo.clicked.connect(self.add)
        self.btRemoverTipo.clicked.connect(self.remover)


    def coboBox(self):
        sql = ("SELECT grupo_atendimento FROM grupo_atendimentos")
        grupo = ConectarSqlite.conectarBd(sql)
        combo = QComboBox()
        for t in grupo:
            for t1 in t:
                combo.addItem(t1)
        return combo

    def carregar(self):
        self.tableEditarServico.setRowCount(0)
        sql2 = ('''SELECT tipo_atendimento, grupo_atendimento, id FROM tipo_atendimentos
        WHERE setor = "''' +login.gsu.tecnico.setor+ '" ORDER BY tipo_atendimento;')
        tipo_atendimento = (ConectarSqlite.conectarBd(sql2))


        for row, tipo in enumerate(tipo_atendimento):
            self.tableEditarServico.insertRow(row)
            for column, item in enumerate(tipo):
                self.tableEditarServico.setItem(row, column, QTableWidgetItem(str(item)))

            combo = self.coboBox()

            combo.setCurrentText(self.tableEditarServico.item(row, 1).text())
            self.tableEditarServico.setCellWidget(row, 1, combo)

    def salvarTipo(self):
        for row in range(self.tableEditarServico.rowCount()):
            tipo = self.tableEditarServico.item(row, 0).text()
            grupo = self.tableEditarServico.cellWidget(row, 1).currentText()
            id = self.tableEditarServico.item(row, 2).text()
            sqlR = ("UPDATE tipo_atendimentos SET tipo_atendimento = '"+ \
            tipo + "', grupo_atendimento = '"+grupo+"' WHERE id = "+id)
            ConectarSqlite.conectarBd(sqlR)

        login.gsu.atualizaComboBoxTipo()
        self.close()

    def add (self):
        row = (self.tableEditarServico.rowCount())
        self.tableEditarServico.insertRow(row)
        self.combo = self.coboBox()
        self.tableEditarServico.setCellWidget(row, 1, self.combo)
        self.combo.setFocus()

        sql = ("SELECT max(id) FROM tipo_atendimentos;")
        result = ConectarSqlite.conectarBd(sql)
        if result[0][0] == None:
            id = 1
        else:
            id = result[0][0] + 1
        tipo = ""
        grupo = self.tableEditarServico.cellWidget(row, 1).currentText()
        setor = login.gsu.tecnico.setor
        sqlR = ("INSERT INTO tipo_atendimentos (id, tipo_atendimento, grupo_atendimento, setor) VALUES ("\
        +str(id)+",'"+str(tipo)+"','"+str(grupo)+"','"+str(setor)+"');")
        ConectarSqlite.conectarBd(sqlR)
        self.tableEditarServico.setItem(row, 2, QTableWidgetItem(str(id)))

    def remover(self):
        rowU = self.tableEditarServico.rowCount()
        row = self.tableEditarServico.currentRow()

        if row > -2:
            id = self.tableEditarServico.item(row, 2).text()
            sql1 = ("DELETE FROM tipo_atendimentos WHERE  id = "+str(id)+";")
            ConectarSqlite.conectarBd(sql1)
            self.tableEditarServico.removeRow(row)
            login.gsu.atualizaComboBoxTipo()

        else:
            id = self.tableEditarServico.item(row, 2).text()
            sql = ("DELETE FROM tipo_atendimentos WHERE  id = "+str(id)+";")
            ConectarSqlite.conectarBd(sql)
            self.tableEditarServico.removeRow(rowU)
            login.gsu.atualizaComboBoxTipo()


now = datetime.datetime.now()
modelo =  r'\\10.7.51.11\ddt$\GSU\01-SUPORT\Controle Atendimentos\PLan\Modelo.xlsx'
modelo2 = r'\\10.7.51.11\ddt$\GSU\01-SUPORT\Controle Atendimentos\PLan\Modelo2.xlsx'


#Carrar janela de login

class Login(QtWidgets.QDialog):

    def __init__(self):
        QtWidgets.QDialog.__init__(self)
        # Carregando Login
        #uic.loadUi("Login.ui", self)

        ############ Inicio Cópia Login.ui ##################
        self.setObjectName("Dialog")
        self.resize(320, 268)
        self.gridLayout = QtWidgets.QGridLayout(self)
        self.gridLayout.setObjectName("gridLayout")
        self.label = QtWidgets.QLabel(self)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 1)
        self.lineLogin = QtWidgets.QLineEdit(self)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineLogin.setFont(font)
        self.lineLogin.setText("")
        self.lineLogin.setObjectName("lineLogin")
        self.gridLayout.addWidget(self.lineLogin, 0, 1, 1, 1)
        self.label_2 = QtWidgets.QLabel(self)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.gridLayout.addWidget(self.label_2, 1, 0, 1, 1)
        self.lineSenha = QtWidgets.QLineEdit(self)
        self.lineSenha.setText("")
        self.lineSenha.setEchoMode(QtWidgets.QLineEdit.Password)
        self.lineSenha.setObjectName("lineSenha")
        self.gridLayout.addWidget(self.lineSenha, 1, 1, 1, 1)
        self.btLogOk = QtWidgets.QDialogButtonBox(self)
        self.btLogOk.setOrientation(QtCore.Qt.Horizontal)
        self.btLogOk.setStandardButtons(QtWidgets.QDialogButtonBox.Ok)
        self.btLogOk.setCenterButtons(True)
        self.btLogOk.setObjectName("btLogOk")
        self.gridLayout.addWidget(self.btLogOk, 2, 0, 1, 3)

        self.retranslateUi(self)
        self.btLogOk.accepted.connect(self.accept)
        self.btLogOk.rejected.connect(self.reject)
        QtCore.QMetaObject.connectSlotsByName(self)

        ###### Ajustes ######
        self.btLogOk.clicked.connect(self.iniciar)

        ###### Ajustes automatizando login com usuário logado #####
        self.username = getpass.getuser() #"6415246"
        self.sql = ("SELECT mat_gm, usuario, nome, setor, senha FROM tecnicos;")
        self.listaTecnicos = ConectarSqlite.conectarBd(self.sql)
        for cont, item in enumerate(self.listaTecnicos):
            if self.username in item[1]:
                self.tecnico = [item]
                self.close()
                return self.login(self.tecnico)

            else:
                self.show()

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "LOGIN"))
        self.label.setText(_translate("Dialog", "Login"))
        self.label_2.setText(_translate("Dialog", "Senha"))
        ############ FiM da Cópia Login.ui ######################

    def iniciar(self):
        self.user = self.lineLogin.text()
        password = self.lineSenha.text()
        sql = ("SELECT mat_gm, senha FROM tecnicos")
        result = ConectarSqlite.conectarBd(sql)
        log = (self.user,str(password))

        if log in result:
            self.sql = ("SELECT mat_gm, usuario, nome, setor, senha FROM tecnicos \
                        WHERE mat_gm = '"+self.user+"';")
            self.tecnico = ConectarSqlite.conectarBd(self.sql)
            login.close()
            return self.login(self.tecnico) #print(self.tecnico)

        else:
            QMessageBox.about(self, 'ERRO', 'Usuário ou senha NÃO CADASTRADO')

    def login(self, item):
        ### Instanciando Formulario atendimentos
        self.gsu = Gsu()

        ### Caregando dados do tecnico logado
        self.gsu.tecnico.setMat(item[0][0])
        self.gsu.tecnico.setUser(item[0][1])
        self.gsu.tecnico.setNome(item[0][2])
        self.gsu.tecnico.setSetor(item[0][3])
        self.gsu.tecnico.setSenha(item[0][4])

        # carregar Matrícula e nome do Tecnico e Setor
        self.gsu.lbSetor.setText(self.gsu.tecnico.setor)
        matT1 = str(self.gsu.tecnico.mat)
        matT = matPadrao(matT1)
        self.gsu.matTecnico.setText(matT)
        nome = nomeT(matT1)
        self.gsu.lbNomeT.setText(nome)

        ## Campo ref a Tipo de atendimento
        self.gsu.atualizaComboBoxTipo()

        #### Contagem de atendimento ####
        self.gsu.contar()

        ### Ajustando formlário para ocultar campos inerente ao DDT_NSS
        if self.gsu.tecnico.setor != 'DDT_NSS':
            self.gsu.tableConsulta.setColumnWidth(5,0)
            self.gsu.chSol_2.hide()
            self.gsu.cbSistema.hide()

        return self.gsu.show()

class Gsu(QMainWindow):

    def __init__(self):
        QMainWindow.__init__(self)
        # Carregando e configurando Formulário
        #uic.loadUi("GSU.ui", self)
        ######## Inicio Cópia #########
        self.setObjectName("gsu")
        self.setEnabled(True)
        self.resize(1462, 691)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Ico.ico"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)
        self.setAutoFillBackground(False)
        self.setStyleSheet("background-image: url(:/newPrefix/logo.jpg);")
        self.layout = QtWidgets.QWidget(self)
        self.layout.setObjectName("layout")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.layout)
        self.gridLayout_4.setHorizontalSpacing(6)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.tabWidget = QtWidgets.QTabWidget(self.layout)
        self.tabWidget.setEnabled(True)
        self.tabWidget.setAcceptDrops(False)
        self.tabWidget.setWhatsThis("")
        self.tabWidget.setAccessibleName("")
        self.tabWidget.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.tabWidget.setStyleSheet("background-image: url(:/newPrefix/logo.jpg);\n"
"image: url(:/newPrefix/logo.jpg);\n"
"border-image: url(:/newPrefix/logo.jpg);")
        self.tabWidget.setTabsClosable(False)
        self.tabWidget.setMovable(False)
        self.tabWidget.setTabBarAutoHide(True)
        self.tabWidget.setObjectName("tabWidget")
        self.Cadastro = QtWidgets.QWidget()
        self.Cadastro.setObjectName("Cadastro")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.Cadastro)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.lbNomeT = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setFamily("Perpetua Titling MT")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.lbNomeT.setFont(font)
        self.lbNomeT.setAlignment(QtCore.Qt.AlignCenter)
        self.lbNomeT.setObjectName("lbNomeT")
        self.gridLayout_2.addWidget(self.lbNomeT, 10, 2, 1, 3)
        self.label_32 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label_32.setFont(font)
        self.label_32.setObjectName("label_32")
        self.gridLayout_2.addWidget(self.label_32, 11, 0, 1, 2)
        self.label_21 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_21.setFont(font)
        self.label_21.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_21.setObjectName("label_21")
        self.gridLayout_2.addWidget(self.label_21, 9, 1, 1, 1)
        self.label_26 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_26.setFont(font)
        self.label_26.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_26.setObjectName("label_26")
        self.gridLayout_2.addWidget(self.label_26, 8, 1, 1, 1)
        self.lbSetor_2 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lbSetor_2.setFont(font)
        self.lbSetor_2.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.lbSetor_2.setObjectName("lbSetor_2")
        self.gridLayout_2.addWidget(self.lbSetor_2, 7, 2, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(637, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem, 6, 0, 1, 4)
        self.comboBoxAtendimento = QtWidgets.QComboBox(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.comboBoxAtendimento.setFont(font)
        self.comboBoxAtendimento.setObjectName("comboBoxAtendimento")
        self.comboBoxAtendimento.addItem("")
        self.comboBoxAtendimento.addItem("")
        self.comboBoxAtendimento.addItem("")
        self.comboBoxAtendimento.addItem("")
        self.comboBoxAtendimento.addItem("")
        self.comboBoxAtendimento.addItem("")
        self.gridLayout_2.addWidget(self.comboBoxAtendimento, 5, 2, 1, 2)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem1, 7, 0, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem2, 9, 4, 1, 1)
        spacerItem3 = QtWidgets.QSpacerItem(75, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem3, 3, 0, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(275, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem4, 1, 4, 1, 1)
        self.label_34 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_34.setFont(font)
        self.label_34.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_34.setObjectName("label_34")
        self.gridLayout_2.addWidget(self.label_34, 0, 1, 1, 1)
        self.label_31 = QtWidgets.QLabel(self.Cadastro)
        self.label_31.setEnabled(False)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_31.setFont(font)
        self.label_31.setObjectName("label_31")
        self.gridLayout_2.addWidget(self.label_31, 1, 6, 1, 1)
        self.dateSolicitacao = QtWidgets.QDateTimeEdit(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(14)
        self.dateSolicitacao.setFont(font)
        self.dateSolicitacao.setObjectName("dateSolicitacao")
        self.gridLayout_2.addWidget(self.dateSolicitacao, 0, 2, 1, 2)
        self.label_28 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_28.setFont(font)
        self.label_28.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_28.setObjectName("label_28")
        self.gridLayout_2.addWidget(self.label_28, 1, 1, 1, 1)
        self.label_23 = QtWidgets.QLabel(self.Cadastro)
        self.label_23.setEnabled(False)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_23.setFont(font)
        self.label_23.setObjectName("label_23")
        self.gridLayout_2.addWidget(self.label_23, 0, 6, 1, 1)
        self.lbDia = QtWidgets.QLabel(self.Cadastro)
        self.lbDia.setEnabled(False)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lbDia.setFont(font)
        self.lbDia.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.lbDia.setObjectName("lbDia")
        self.gridLayout_2.addWidget(self.lbDia, 0, 5, 1, 1)
        self.chSol = QtWidgets.QCheckBox(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.chSol.setFont(font)
        self.chSol.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.chSol.setObjectName("chSol")
        self.gridLayout_2.addWidget(self.chSol, 4, 1, 1, 1)
        self.dateConcluido = QtWidgets.QDateTimeEdit(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(14)
        self.dateConcluido.setFont(font)
        self.dateConcluido.setObjectName("dateConcluido")
        self.gridLayout_2.addWidget(self.dateConcluido, 1, 2, 1, 2)
        self.lbMes = QtWidgets.QLabel(self.Cadastro)
        self.lbMes.setEnabled(False)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lbMes.setFont(font)
        self.lbMes.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.lbMes.setObjectName("lbMes")
        self.gridLayout_2.addWidget(self.lbMes, 1, 5, 1, 1)
        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem5, 4, 9, 1, 1)
        self.checkBox_2 = QtWidgets.QCheckBox(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(14)
        self.checkBox_2.setFont(font)
        self.checkBox_2.setChecked(False)
        self.checkBox_2.setTristate(False)
        self.checkBox_2.setObjectName("checkBox_2")
        self.gridLayout_2.addWidget(self.checkBox_2, 3, 8, 1, 1)
        self.lineSol = QtWidgets.QLineEdit(self.Cadastro)
        self.lineSol.setEnabled(False)
        font = QtGui.QFont()
        font.setPointSize(11)
        self.lineSol.setFont(font)
        self.lineSol.setText("")
        self.lineSol.setObjectName("lineSol")
        self.gridLayout_2.addWidget(self.lineSol, 4, 2, 1, 2)
        self.chSol_2 = QtWidgets.QCheckBox(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.chSol_2.setFont(font)
        self.chSol_2.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.chSol_2.setObjectName("chSol_2")
        self.gridLayout_2.addWidget(self.chSol_2, 3, 1, 1, 1)
        self.label_3 = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_3.setObjectName("label_3")
        self.gridLayout_2.addWidget(self.label_3, 5, 1, 1, 1)
        self.lbSetor = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lbSetor.setFont(font)
        self.lbSetor.setText("")
        self.lbSetor.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignVCenter)
        self.lbSetor.setObjectName("lbSetor")
        self.gridLayout_2.addWidget(self.lbSetor, 7, 3, 1, 1)
        self.widgetVnc_2 = QtWidgets.QWidget(self.Cadastro)
        self.widgetVnc_2.setEnabled(False)
        self.widgetVnc_2.setObjectName("widgetVnc_2")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.widgetVnc_2)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.label_36 = QtWidgets.QLabel(self.widgetVnc_2)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_36.setFont(font)
        self.label_36.setObjectName("label_36")
        self.gridLayout_6.addWidget(self.label_36, 0, 0, 1, 2)
        self.timeVpnInicio = QtWidgets.QTimeEdit(self.widgetVnc_2)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setStrikeOut(False)
        self.timeVpnInicio.setFont(font)
        self.timeVpnInicio.setAcceptDrops(False)
        self.timeVpnInicio.setAlignment(QtCore.Qt.AlignCenter)
        self.timeVpnInicio.setObjectName("timeVpnInicio")
        self.gridLayout_6.addWidget(self.timeVpnInicio, 0, 2, 1, 1)
        self.label_35 = QtWidgets.QLabel(self.widgetVnc_2)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_35.setFont(font)
        self.label_35.setObjectName("label_35")
        self.gridLayout_6.addWidget(self.label_35, 1, 0, 1, 2)
        self.timeVpnTermino = QtWidgets.QTimeEdit(self.widgetVnc_2)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setStrikeOut(False)
        self.timeVpnTermino.setFont(font)
        self.timeVpnTermino.setAcceptDrops(False)
        self.timeVpnTermino.setAlignment(QtCore.Qt.AlignCenter)
        self.timeVpnTermino.setObjectName("timeVpnTermino")
        self.gridLayout_6.addWidget(self.timeVpnTermino, 1, 2, 1, 1)
        self.label_2 = QtWidgets.QLabel(self.widgetVnc_2)
        self.label_2.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_2.setObjectName("label_2")
        self.gridLayout_6.addWidget(self.label_2, 2, 0, 1, 1)
        self.ip = QtWidgets.QLineEdit(self.widgetVnc_2)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.ip.setFont(font)
        self.ip.setObjectName("ip")
        self.gridLayout_6.addWidget(self.ip, 2, 1, 1, 2)
        self.gridLayout_2.addWidget(self.widgetVnc_2, 4, 7, 3, 2)
        self.salvar = QtWidgets.QPushButton(self.Cadastro)
        self.salvar.setObjectName("salvar")
        self.gridLayout_2.addWidget(self.salvar, 14, 5, 1, 1)
        spacerItem6 = QtWidgets.QSpacerItem(661, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem6, 11, 2, 1, 3)
        self.btLimpar = QtWidgets.QPushButton(self.Cadastro)
        self.btLimpar.setObjectName("btLimpar")
        self.gridLayout_2.addWidget(self.btLimpar, 14, 6, 1, 1)
        spacerItem7 = QtWidgets.QSpacerItem(445, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem7, 12, 0, 1, 2)
        spacerItem8 = QtWidgets.QSpacerItem(138, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem8, 14, 8, 1, 1)
        self.cancelar = QtWidgets.QPushButton(self.Cadastro)
        self.cancelar.setObjectName("cancelar")
        self.gridLayout_2.addWidget(self.cancelar, 14, 7, 1, 1)
        self.line = QtWidgets.QFrame(self.Cadastro)
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.gridLayout_2.addWidget(self.line, 2, 0, 1, 9)
        self.cbSistema = QtWidgets.QComboBox(self.Cadastro)
        self.cbSistema.setEnabled(False)
        self.cbSistema.setObjectName("cbSistema")
        self.cbSistema.addItem("")
        self.cbSistema.setItemText(0, "")
        self.gridLayout_2.addWidget(self.cbSistema, 3, 2, 1, 2)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.tipoServico = QtWidgets.QComboBox(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.tipoServico.setFont(font)
        self.tipoServico.setObjectName("tipoServico")
        self.horizontalLayout.addWidget(self.tipoServico)
        self.toolTipo = QtWidgets.QToolButton(self.Cadastro)
        self.toolTipo.setObjectName("toolTipo")
        self.horizontalLayout.addWidget(self.toolTipo)
        self.gridLayout_2.addLayout(self.horizontalLayout, 8, 2, 1, 3)
        self.observacao = QtWidgets.QTextEdit(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.observacao.setFont(font)
        self.observacao.setObjectName("observacao")
        self.gridLayout_2.addWidget(self.observacao, 13, 0, 1, 9)
        self.groupBox = QtWidgets.QGroupBox(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.groupBox)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.label_30 = QtWidgets.QLabel(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.label_30.setFont(font)
        self.label_30.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_30.setObjectName("label_30")
        self.gridLayout_3.addWidget(self.label_30, 0, 0, 1, 1)
        self.matSolicitante = QtWidgets.QLineEdit(self.groupBox)
        self.matSolicitante.setEnabled(True)
        font = QtGui.QFont()
        font.setPointSize(11)
        self.matSolicitante.setFont(font)
        self.matSolicitante.setText("")
        self.matSolicitante.setObjectName("matSolicitante")
        self.gridLayout_3.addWidget(self.matSolicitante, 0, 1, 1, 2)
        self.label_33 = QtWidgets.QLabel(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.label_33.setFont(font)
        self.label_33.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_33.setObjectName("label_33")
        self.gridLayout_3.addWidget(self.label_33, 2, 0, 1, 2)
        self.PesqUnidade = QtWidgets.QLineEdit(self.groupBox)
        self.PesqUnidade.setEnabled(True)
        font = QtGui.QFont()
        font.setPointSize(11)
        self.PesqUnidade.setFont(font)
        self.PesqUnidade.setText("")
        self.PesqUnidade.setObjectName("PesqUnidade")
        self.gridLayout_3.addWidget(self.PesqUnidade, 2, 2, 1, 1)
        self.label = QtWidgets.QLabel(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label.setObjectName("label")
        self.gridLayout_3.addWidget(self.label, 3, 0, 1, 1)
        self.unidadeSolicitante = QtWidgets.QComboBox(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(9)
        self.unidadeSolicitante.setFont(font)
        self.unidadeSolicitante.setObjectName("unidadeSolicitante")
        self.gridLayout_3.addWidget(self.unidadeSolicitante, 3, 1, 1, 2)
        self.lbNomeC = QtWidgets.QLabel(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.lbNomeC.setFont(font)
        self.lbNomeC.setAlignment(QtCore.Qt.AlignCenter)
        self.lbNomeC.setObjectName("lbNomeC")
        self.gridLayout_3.addWidget(self.lbNomeC, 1, 0, 1, 3)
        self.gridLayout_2.addWidget(self.groupBox, 7, 5, 6, 4)
        self.checkBox_data = QtWidgets.QCheckBox(self.Cadastro)
        self.checkBox_data.setChecked(True)
        self.checkBox_data.setObjectName("checkBox_data")
        self.gridLayout_2.addWidget(self.checkBox_data, 0, 4, 1, 1)
        self.matTecnico = QtWidgets.QLabel(self.Cadastro)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.matTecnico.setFont(font)
        self.matTecnico.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.matTecnico.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignVCenter)
        self.matTecnico.setObjectName("matTecnico")
        self.gridLayout_2.addWidget(self.matTecnico, 9, 2, 1, 2)
        self.bt_agenda = QtWidgets.QPushButton(self.Cadastro)
        self.bt_agenda.setObjectName("bt_agenda")
        self.gridLayout_2.addWidget(self.bt_agenda, 14, 0, 1, 1)
        self.tabWidget.addTab(self.Cadastro, "")
        self.Consulta = QtWidgets.QWidget()
        self.Consulta.setObjectName("Consulta")
        self.gridLayout = QtWidgets.QGridLayout(self.Consulta)
        self.gridLayout.setObjectName("gridLayout")
        self.btGerarRelatorio = QtWidgets.QPushButton(self.Consulta)
        self.btGerarRelatorio.setEnabled(True)
        self.btGerarRelatorio.setObjectName("btGerarRelatorio")
        self.gridLayout.addWidget(self.btGerarRelatorio, 2, 0, 1, 1)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.gridLayout.addLayout(self.horizontalLayout_4, 0, 6, 1, 1)
        self.btConsultar = QtWidgets.QPushButton(self.Consulta)
        self.btConsultar.setObjectName("btConsultar")
        self.gridLayout.addWidget(self.btConsultar, 0, 5, 1, 1)
        self.fechar = QtWidgets.QPushButton(self.Consulta)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.fechar.setFont(font)
        self.fechar.setObjectName("fechar")
        self.gridLayout.addWidget(self.fechar, 3, 5, 1, 1)
        spacerItem9 = QtWidgets.QSpacerItem(672, 23, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem9, 3, 6, 1, 1)
        self.scrollArea = QtWidgets.QScrollArea(self.Consulta)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents_2 = QtWidgets.QWidget()
        self.scrollAreaWidgetContents_2.setGeometry(QtCore.QRect(0, 0, 1422, 520))
        self.scrollAreaWidgetContents_2.setObjectName("scrollAreaWidgetContents_2")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents_2)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.tableConsulta = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
        self.tableConsulta.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableConsulta.setShowGrid(True)
        self.tableConsulta.setRowCount(0)
        self.tableConsulta.setColumnCount(16)
        self.tableConsulta.setObjectName("tableConsulta")
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignJustify|QtCore.Qt.AlignVCenter)
        self.tableConsulta.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.tableConsulta.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(13, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(14, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableConsulta.setHorizontalHeaderItem(15, item)
        self.gridLayout_5.addWidget(self.tableConsulta, 1, 0, 1, 1)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents_2)
        self.gridLayout.addWidget(self.scrollArea, 1, 0, 1, 7)
        self.dateInicioPeriodo = QtWidgets.QDateEdit(self.Consulta)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.dateInicioPeriodo.setFont(font)
        self.dateInicioPeriodo.setObjectName("dateInicioPeriodo")
        self.gridLayout.addWidget(self.dateInicioPeriodo, 0, 2, 1, 1)
        self.label_38 = QtWidgets.QLabel(self.Consulta)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_38.setFont(font)
        self.label_38.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.label_38.setAutoFillBackground(False)
        self.label_38.setAlignment(QtCore.Qt.AlignCenter)
        self.label_38.setObjectName("label_38")
        self.gridLayout.addWidget(self.label_38, 0, 3, 1, 1)
        self.label_37 = QtWidgets.QLabel(self.Consulta)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_37.setFont(font)
        self.label_37.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.label_37.setTextFormat(QtCore.Qt.RichText)
        self.label_37.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_37.setObjectName("label_37")
        self.gridLayout.addWidget(self.label_37, 0, 1, 1, 1)
        self.dateFimPeriodo = QtWidgets.QDateEdit(self.Consulta)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.dateFimPeriodo.setFont(font)
        self.dateFimPeriodo.setObjectName("dateFimPeriodo")
        self.gridLayout.addWidget(self.dateFimPeriodo, 0, 4, 1, 1)
        self.btEditar = QtWidgets.QPushButton(self.Consulta)
        self.btEditar.setEnabled(True)
        self.btEditar.setObjectName("btEditar")
        self.gridLayout.addWidget(self.btEditar, 2, 1, 1, 1)
        self.unificado = QtWidgets.QPushButton(self.Consulta)
        self.unificado.setEnabled(True)
        self.unificado.setObjectName("unificado")
        self.gridLayout.addWidget(self.unificado, 3, 0, 1, 1)
        self.btConsultar.raise_()
        self.btGerarRelatorio.raise_()
        self.scrollArea.raise_()
        self.fechar.raise_()
        self.label_37.raise_()
        self.dateInicioPeriodo.raise_()
        self.label_38.raise_()
        self.dateFimPeriodo.raise_()
        self.btEditar.raise_()
        self.unificado.raise_()
        self.tabWidget.addTab(self.Consulta, "")
        self.gridLayout_4.addWidget(self.tabWidget, 0, 1, 1, 1)
        self.setCentralWidget(self.layout)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)

        self.retranslateUi(self)
        self.tabWidget.setCurrentIndex(0)
        self.cancelar.clicked.connect(self.close)
        self.chSol.clicked['bool'].connect(self.lineSol.setEnabled)
        self.chSol_2.clicked['bool'].connect(self.cbSistema.setEnabled)
        self.fechar.clicked.connect(self.close)
        self.checkBox_2.clicked['bool'].connect(self.widgetVnc_2.setEnabled)
        QtCore.QMetaObject.connectSlotsByName(self)

    def retranslateUi(self, gsu):
        _translate = QtCore.QCoreApplication.translate
        gsu.setWindowTitle(_translate("gsu", "DDT_GSU    v2.0.0"))
        self.lbNomeT.setText(_translate("gsu", "NOME"))
        self.label_32.setText(_translate("gsu", "Descrição do atendimento"))
        self.label_21.setText(_translate("gsu", "Matrícula do Técnico:"))
        self.label_26.setText(_translate("gsu", "Tipo de serviço:"))
        self.lbSetor_2.setText(_translate("gsu", "SETOR:"))
        self.comboBoxAtendimento.setItemText(0, _translate("gsu", "SOL"))
        self.comboBoxAtendimento.setItemText(1, _translate("gsu", "Presencial"))
        self.comboBoxAtendimento.setItemText(2, _translate("gsu", "Visita Técinca"))
        self.comboBoxAtendimento.setItemText(3, _translate("gsu", "Telefone"))
        self.comboBoxAtendimento.setItemText(4, _translate("gsu", "Memorando"))
        self.comboBoxAtendimento.setItemText(5, _translate("gsu", "Email"))
        self.label_34.setText(_translate("gsu", "Data da Solicitação:"))
        self.label_31.setText(_translate("gsu", "Atendidos no Mês"))
        self.label_28.setText(_translate("gsu", "Serviço Concluido Em:"))
        self.label_23.setText(_translate("gsu", "Atendidos no Dia"))
        self.lbDia.setText(_translate("gsu", "0"))
        self.chSol.setText(_translate("gsu", "      Atendimento com chamado no sol Nº:  "))
        self.lbMes.setText(_translate("gsu", "0"))
        self.checkBox_2.setText(_translate("gsu", "  Feito acesso remoto"))
        self.chSol_2.setText(_translate("gsu", "                               Suporte ao  Sitema:"))
        self.label_3.setText(_translate("gsu", "Forma de Contato:  "))
        self.label_36.setText(_translate("gsu", "Horário de Início do acesso:"))
        self.label_35.setText(_translate("gsu", "Témino do Acesso:"))
        self.label_2.setText(_translate("gsu", "IP ACESSADO:"))
        self.salvar.setText(_translate("gsu", "Salvar"))
        self.btLimpar.setText(_translate("gsu", "Limpar Campos"))
        self.cancelar.setText(_translate("gsu", "Cancelar"))
        self.toolTipo.setText(_translate("gsu", "..."))
        self.observacao.setHtml(_translate("gsu", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:12pt; font-weight:400; font-style:normal;\">\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>"))
        self.groupBox.setTitle(_translate("gsu", "Solicitante:"))
        self.label_30.setText(_translate("gsu", "Matrícula:"))
        self.label_33.setText(_translate("gsu", "Pesquisar Unidade:"))
        self.label.setText(_translate("gsu", "Unidade"))
        self.lbNomeC.setText(_translate("gsu", ""))
        self.checkBox_data.setText(_translate("gsu", "Atualizar data após salvar"))
        self.matTecnico.setText(_translate("gsu", "TextLabel"))
        self.bt_agenda.setText(_translate("gsu", "Agenda Telefônica"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.Cadastro), _translate("gsu", "Cadastro"))
        self.btGerarRelatorio.setText(_translate("gsu", "Gerar Relatório do SETOR"))
        self.btConsultar.setText(_translate("gsu", "Atualizar Consulta"))
        self.fechar.setText(_translate("gsu", "Fechar"))
        item = self.tableConsulta.horizontalHeaderItem(0)
        item.setText(_translate("gsu", "ID"))
        item = self.tableConsulta.horizontalHeaderItem(1)
        item.setText(_translate("gsu", "TÉCNICO"))
        item = self.tableConsulta.horizontalHeaderItem(2)
        item.setText(_translate("gsu", "SOLICITANTE \n"
"  MATRÍCULA"))
        item = self.tableConsulta.horizontalHeaderItem(3)
        item.setText(_translate("gsu", "SOLICITANTE \n"
"     NOME"))
        item = self.tableConsulta.horizontalHeaderItem(4)
        item.setText(_translate("gsu", "UNIDADE"))
        item = self.tableConsulta.horizontalHeaderItem(5)
        item.setText(_translate("gsu", "Suporte ao Sistema"))
        item = self.tableConsulta.horizontalHeaderItem(6)
        item.setText(_translate("gsu", "Número do SOL"))
        item = self.tableConsulta.horizontalHeaderItem(7)
        item.setText(_translate("gsu", "    FORMA \n"
"DE CONTATO"))
        item = self.tableConsulta.horizontalHeaderItem(8)
        item.setText(_translate("gsu", "TIPO DE SERVIÇO"))
        item = self.tableConsulta.horizontalHeaderItem(9)
        item.setText(_translate("gsu", "  DATA/HORA  \n"
"DA SOLICITAÇÃO"))
        item = self.tableConsulta.horizontalHeaderItem(10)
        item.setText(_translate("gsu", "DATA/HORA \n"
" Finalizado"))
        item = self.tableConsulta.horizontalHeaderItem(11)
        item.setText(_translate("gsu", "Descrição do Serviço"))
        item = self.tableConsulta.horizontalHeaderItem(12)
        item.setText(_translate("gsu", "Acesso Remoto"))
        item = self.tableConsulta.horizontalHeaderItem(13)
        item.setText(_translate("gsu", "INICIO do Acesso"))
        item = self.tableConsulta.horizontalHeaderItem(14)
        item.setText(_translate("gsu", "Fim do acesso"))
        item = self.tableConsulta.horizontalHeaderItem(15)
        item.setText(_translate("gsu", "IP do PC"))
        self.label_38.setText(_translate("gsu", "à"))
        self.label_37.setText(_translate("gsu", "De"))
        self.btEditar.setText(_translate("gsu", "Editar Linha Atual"))
        self.unificado.setText(_translate("gsu", "Relatório Unificado"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.Consulta), _translate("gsu", "Consulta"))
############## Fim Cópia #############

        self.tecnico = Tecnico("","","","t","")
        self.dateSolicitacao.setDateTime(now)
        self.dateConcluido.setDateTime(datetime.datetime.now() + datetime.timedelta(minutes=5))
        self.dateInicioPeriodo.setDate(now)
        self.dateFimPeriodo.setDate(now)

        #### Configurando aba consulta ####
        self.btConsultar.clicked.connect(self.atualizaConsulta)
        self.btGerarRelatorio.clicked.connect(self.relatorio)
        self.unificado.clicked.connect(self.relatorioUnificado)
        self.btEditar.clicked.connect(self.bt_editar)
        self.bt_agenda.clicked.connect(self.agenda)

        ### Ajustando células da tabela consulta
        self.tableConsulta.setColumnWidth(0,0)
        self.tableConsulta.setColumnWidth(1, 250)
        self.tableConsulta.setColumnWidth(2, 250)
        self.tableConsulta.setColumnWidth(3, 250)
        self.tableConsulta.setColumnWidth(4, 250)
        self.tableConsulta.setColumnWidth(5, 250)
        self.tableConsulta.setColumnWidth(6, 250)
        self.tableConsulta.setColumnWidth(7, 250)
        self.tableConsulta.setColumnWidth(8, 150)
        self.tableConsulta.setColumnWidth(9, 150)
        self.tableConsulta.setColumnWidth(10, 150)
        self.tableConsulta.setColumnWidth(14, 150)

        ###### Campos do formulário de atendimento #######
        ## Campos ref a Unidades
        self.PesqUnidade.textChanged.connect(self.locCbUnidade)
        Gsu.atualizaComboBoxUnidade(self)
        Gsu.atualizaComboBoxSistema(self)
        self.matSolicitante.editingFinished.connect(self.nomeC)

        ###### Fim dos campos ref ao formulário atendimento #######
        self.toolTipo.clicked.connect(self.editarTipoAtendimento)

        #### Funcionalidades dos butãos do Formulário
        self.salvar.clicked.connect(self.salvo)



    def agenda(self):
        self.agenda1 = agenda.Agenda()
        self.agenda1.show()
    #### Localiza e filtra as unidades
    def locCbUnidade (self):
        sql = ("SELECT id, unidade FROM unidades ORDER BY unidade;")
        result = (ConectarSqlite.conectarBd(sql))
        texto = self.PesqUnidade.text()
        texto = texto.upper()
        list = []

        try:
            for cont, item in enumerate(result):
                if item[1][0:2] == 'GM':
                    if  item[1].count(texto):
                        list.append(item[1])
            self.unidadeSolicitante.setCurrentText(list[0])
        except:
            pass

    #### Carregando as ComboBox das unidades
    def atualizaComboBoxUnidade(self):
        self.unidadeSolicitante.clear()
        sql = ("SELECT id, unidade FROM unidades ORDER BY unidade;")
        result = (ConectarSqlite.conectarBd(sql))
        for cont, item in enumerate(result):
            #print(item[1][0:2])
            if item[1][0:2] == 'GM':
                self.unidadeSolicitante.insertItem(item[0],str(item[1]))

    def atualizaComboBoxSistema(self):
        self.cbSistema.clear()
        sql = ("SELECT id, sistema FROM sistemas ORDER BY sistema;")
        result = (ConectarSqlite.conectarBd(sql))
        for cont, item in enumerate(result):
            self.cbSistema.insertItem(item[0],str(item[1]))

    def atualizaComboBoxTipo(self):
        self.tipoServico.clear()
        sql = ("SELECT id, tipo_atendimento FROM tipo_atendimentos WHERE setor = '"+self.tecnico.setor+"' ORDER BY tipo_atendimento;")
        result = ConectarSqlite.conectarBd(sql)
        for cont, item in enumerate(result):
            self.tipoServico.insertItem(item[0],str(item[1]))

    def editarTipoAtendimento(self):
        self.tipoAtendimento = EditarTipoAtendimento()
        self.tipoAtendimento.carregar()
        self.tipoAtendimento.show()


    def nomeC(self):
        mat = self.matSolicitante.text()
        matC = matPadrao(mat)
        self.matSolicitante.setText(matC)


    def salvo(self):
        Gsu.contar(self)

        # Obtendo dados do Formulario
        a = self.lbSetor.text()
        b = self.cbSistema.currentText()
        c = self.matTecnico.text()
        d = self.matSolicitante.text()
        e = self.unidadeSolicitante.currentText()
        f = self.lineSol.text()
        g = self.comboBoxAtendimento.currentText()
        h = self.tipoServico.currentText()
        #### Pegar Grupo
        sqlG = ("SELECT grupo_atendimento FROM tipo_atendimentos WHERE setor = '"+a+"' AND tipo_atendimento LIKE '%"+h+"%';")
        resltG = ConectarSqlite.conectarBd(sqlG)
        h1 = resltG[0][0]
        i = self.dateSolicitacao.dateTime().toString('yyyy-MM-dd hh:mm')
        j = self.dateConcluido.dateTime().toString('yyyy-MM-dd hh:mm')
        k = self.observacao.toPlainText()

        if self.checkBox_2.isChecked():
            l = 'SIM'
            m = self.timeVpnInicio.time().toString('hh:mm')
            n = self.timeVpnTermino.time().toString('hh:mm')
            o = self.ip.text()
        else:
            l = 'Não'
            m = "00:00"
            n = "00:00"
            o = "Sem acesso Remoto"

        nomeT = self.lbNomeT.text()
        nomeC = self.lbNomeC.text()

        # Salvando dados no SQLITE
        sql1 = ("SELECT max(id) FROM atendimentos;")
        result = ConectarSqlite.conectarBd(sql1)
        if result[0][0] == None:
            id = 1
        else:
            id = result[0][0] + 1
        tupla = (id,a,c,d,e,b,f,g,h,h1,i,j,k,l,m,n,o)
        sql = ('''INSERT INTO atendimentos(id,setor, mat_tecnico, mat_solicitante, unidade, sistema, sol, forma_contato, tipo_servico, grupo_servico,solicitacao_inicio, solicitacao_fim, descricao, remoto, hInicio, hFim, ip)
        VALUES ''' + str(tupla))
        ConectarSqlite.conectarBd(sql)
        QMessageBox.about(self, 'Informação', 'Cadastro Salvo')

        Gsu.limpar(self)
        Gsu.contar(self)

        if self.checkBox_data.isChecked():
            self.dateSolicitacao.setDateTime(datetime.datetime.now())
            self.dateConcluido.setDateTime(datetime.datetime.now()+datetime.timedelta(minutes=5))

    def contar(self):
        setor = self.tecnico.setor
        hj = QDate.currentDate()
        hoje = hj.currentDate().toString('yyyy-MM-dd')
        hoje1 = hj.currentDate().toString('yyyy-MM')
        sql = ("SELECT COUNT(solicitacao_fim) FROM atendimentos WHERE setor = '"+setor+"' AND solicitacao_fim LIKE'%" + hoje + "%';")
        sql1 = ("SELECT COUNT(solicitacao_fim) FROM atendimentos WHERE setor = '"+setor+"' AND solicitacao_fim LIKE'%" + hoje1 + "%';")
        nome = (ConectarSqlite.conectarBd(sql))
        nome1 = (ConectarSqlite.conectarBd(sql1))
        dia = (nome[0][0])
        mes = (nome1[0][0])

        self.lbDia.setText(str(dia))
        self.lbMes.setText(str(mes))
        self.tableConsulta.setRowCount(mes)

    def atualizaConsulta(self):
        Gsu.contar(self)

        setor = self.lbSetor.text()
        inicio = self.dateInicioPeriodo.date()
        inicio = inicio.toString('yyyy-MM-dd')
        fim = self.dateFimPeriodo.date()
        fim = fim.addDays(1)
        fim = fim.toString('yyyy-MM-dd')
        self.tableConsulta.setRowCount(0)

        sql2 = ("SELECT id,mat_tecnico,mat_solicitante,mat_solicitante,unidade,sistema,sol,forma_contato,tipo_servico,solicitacao_inicio,solicitacao_fim,descricao,remoto,hInicio,hFim,ip"
                " FROM atendimentos WHERE setor = '"+setor+"' AND solicitacao_fim BETWEEN '" + inicio + "' AND '" + fim + "' ORDER BY solicitacao_fim;")
        result = (ConectarSqlite.conectarBd(sql2))


        for row, form in enumerate(result):
            try:
                matT = form[1][0:3]+form[1][4:7]+form[1][8]
            except:
                matT = form[1]

            sqlT=("SELECT nome FROM tecnicos WHERE mat_gm LIKE '%"+str(matT)+"%';")
            sqlC=("SELECT nome FROM servidores WHERE mat_gm1 = '"+form[3]+"';")
            resultC = (ConectarSqlite.conectarBd(sqlC))
            resultT = (ConectarSqlite.conectarBd(sqlT))
            form = list(form)

            try:
                form[1]=resultT[0][0]
            except:
                pass
            try:
                resultC = str(resultC[0][0])
                form[3]=resultC.upper()
            except:
                pass

            self.tableConsulta.insertRow(row)
            for column, item in enumerate(form):
                self.tableConsulta.setItem(row, column, QTableWidgetItem(str(item)))

        if self.checkBox_data.isChecked():
            self.dateSolicitacao.setDateTime(datetime.datetime.now())
            self.dateConcluido.setDateTime(datetime.datetime.now()+datetime.timedelta(minutes=5))

    def relatorio(self):
        ##Altenticando
        matT = self.tecnico.mat
        if acesso(matT,"0"):
            pass
        else:
            QMessageBox.about(self, 'Erro 0', 'Acesso negado')
            return
        ##Corpo da função
        setor = self.lbSetor.text()
        inicio = self.dateInicioPeriodo.date().toString('yyyy-MM-dd')
        fim = self.dateFimPeriodo.date()
        fim = fim.addDays(1)
        fim = fim.toString('yyyy-MM-dd')

        # Preparando Planilha
        wb2 = openpyxl.load_workbook(modelo)
        sh2 = wb2.get_sheet_by_name("Atendimentos")
        ws2 = wb2.active
        # Limpar dados
        for row in range(2, ws2.max_row + 1):
            for column in range(1, ws2.max_column + 1):
                ws2.cell(row=row, column=column).value = None

        # Obitendo dados do Banco de dados
        wb2.save(modelo)

        sql2 = ("SELECT setor,mat_tecnico,mat_solicitante,mat_solicitante,unidade,sistema,sol,forma_contato,tipo_servico,solicitacao_inicio,solicitacao_fim,descricao,remoto,hInicio,hFim,ip"
                " FROM atendimentos WHERE setor = '"+setor+"' AND solicitacao_fim BETWEEN '" + inicio + "' AND '" + fim + "' ORDER BY solicitacao_fim;")
        result = (ConectarSqlite.conectarBd(sql2))

        ln = 1
        #Incrementando lista gerada pelo banco de dados
        for row, form in enumerate(result):
            try:
                sqlT=("SELECT nome FROM servidores WHERE mat_gm1 = '"+form[1]+"';")
                sqlC=("SELECT nome FROM servidores WHERE mat_gm1 = '"+form[3]+"';")
                resultC = (ConectarSqlite.conectarBd(sqlC))
                resultT = (ConectarSqlite.conectarBd(sqlT))
                lista = list(form)
                lista[3]=resultC[0][0]
                lista[1]=resultT[0][0]
            except:
                pass

            ln = ln + 1
            k = "=L"+str(ln)+"-K"+str(ln)

            lista = [row + 1] + lista
            lista.insert(12, k)

            # Passando os dados para Planilha
            for column, item in enumerate(lista):
                ws2.cell(row=row + 2, column=column + 1).value=item
                ws2.cell(row=row + 2, column=column + 1).alignment=openpyxl.styles.Alignment(vertical='center', horizontal='center')
        wb2.save(modelo)
        # Salvando relatorio no diretório escolhido pelo usuário
        try:
            # if user != 'DDT_NSS':
            #     sh2.column_dimensions['G'].hidden= True
            filename = QFileDialog.getSaveFileName(self, 'SALVAR', os.getenv('USERPROFILE'),'Arquivo (*.xlsx);;All Files (*)')
            wb2.save(filename[0])
            os.startfile(filename[0])
        except:
            pass

    def relatorioUnificado(self):
        ##Autentificação
        matT = self.tecnico.mat
        if acesso(matT,"01"):
            pass
        else:
            QMessageBox.about(self, 'Erro 01', 'Acesso negado')
            return
        ##Corpo da função
        inicio = self.dateInicioPeriodo.date().toString('yyyy-MM-dd')
        fim = self.dateFimPeriodo.date()
        fim = fim.addDays(1)
        fim = fim.toString('yyyy-MM-dd')

        # Preparando Planilha
        wb2 = openpyxl.load_workbook(modelo2)
        sh2 = wb2.get_sheet_by_name("Atendimentos")
        ws2 = wb2.active
        # Limpar dados
        for row in range(2, ws2.max_row + 1):
            for column in range(1, ws2.max_column + 1):
                ws2.cell(row=row, column=column).value = None
        wb2.save(modelo2)
        # Obitendo dados do Banco de dados
        sql = (
                "SELECT  a.setor, b.grupo, a.forma_contato, a.grupo_servico, a.solicitacao_inicio, a.solicitacao_fim FROM "
                "atendimentos a INNER JOIN unidades b on b.unidade = a.unidade"
                " WHERE a.solicitacao_fim BETWEEN '" + inicio + "' AND '" + fim + "' order by a.solicitacao_fim")
        result = (ConectarSqlite.conectarBd(sql))

        ln = 1
        # Incrementando lista gerada pelo banco de dados
        for row, form in enumerate(result):
            lista = (list(form))
            lista = [row + 1] + lista
            ln = ln + 1
            k = "=G"+str(ln)+"-F"+str(ln)
            lista.insert(7, k)
            inicio = lista[5].split(" ")
            fim = lista[6].split(" ")
            lista[5] = inicio[0]
            lista[6] = inicio[0]
            print(lista)

            # Passando os dados para Planilha
            for column, item in enumerate(lista):
                ws2.cell(row=row + 2, column=column + 1).value=item
                ws2.cell(row=row + 2, column=column + 1).alignment=openpyxl.styles.Alignment(vertical='center', horizontal='center')
        wb2.save(modelo2)

        # Salvando relatorio no diretório escolhido pelo usuário
        try:
            filename = QFileDialog.getSaveFileName(self, 'SALVAR', os.getenv('USERPROFILE'),'Arquivo (*.xlsx);;All Files (*)')
            wb2.save(filename[0])
            os.startfile(filename[0])
        except:
            pass

    def limpar(self):
        self.matSolicitante.clear()
        self.lineSol.clear()
        self.observacao.clear()
        self.checkBox_2.setChecked(False)
        self.widgetVnc_2.setEnabled(False)
        self.lbNomeC.setText("")

    def bt_editar(self):
        ## Altenticando
        matT = self.tecnico.mat
        if acesso(matT,"02"):
            pass
        else:
            QMessageBox.about(self, 'Erro 02', 'Acesso negado! \n\n\nNecessita da permissão: "02"')
            return
        ##Corpo da função
        row = self.tableConsulta.currentRow()
        self.edit = Editar()
        self.edit.id_editar = int(self.tableConsulta.item(row,0).text())
        self.edit.carregar()





if __name__ == "__main__":

    app = QApplication(sys.argv)
    login = Login()
    app.exec_()
